from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4

def convert_excel_to_pdf(input_path, output_path):
    wb = load_workbook(input_path)
    sheet = wb.active
    c = canvas.Canvas(output_path, pagesize=A4)
    width, height = A4
    y = height - 50
    for row in sheet.iter_rows(values_only=True):
        line = '\t'.join([str(cell) if cell is not None else '' for cell in row])
        c.drawString(50, y, line)
        y -= 15
        if y < 50:
            c.showPage()
            y = height - 50
    c.save()
